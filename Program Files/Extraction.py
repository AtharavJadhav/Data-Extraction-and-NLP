from cmath import polar
from bs4 import BeautifulSoup
import requests
import openpyxl
import os.path
import nltk
import re
import string 
import csv

#nltk.download('all')
#nltk.download('wordnet')
#nltk.download('stopwords')

from nltk.corpus import stopwords
from pyphen import Pyphen

p = Pyphen(lang='en_US')

wb = openpyxl.load_workbook('Input.xlsx')
ws = wb['Sheet1']

i = 2
j = ws.cell(row=2, column=1).value
n = 5
while i<=n:

     try:
        a = ws.cell(row=i, column=2).hyperlink.target
     except:
        a = ws.cell(row=i, column=2).value
        
     print(a)
     i = i+1
     req = requests.get(a)
     save_path = 'C:/Users/Atharav Jadhav/source/repos/PythonApplication1/TextFiles/'
     soup = BeautifulSoup(req.content, "html.parser")
     cnv_j = str(j)
     completeName = os.path.join(save_path, cnv_j)
     f = open(completeName+".txt", "w")
     #traverse paragraphs from soup
     for data in soup.find_all("p"):
        x = " "
        sum = data.get_text()
        f.writelines(x)
        f.writelines(sum)
     f.close()
     j = j+1

#Till here, all extraction part is over
directory = 'C:/Users/Atharav Jadhav/source/repos/PythonApplication1/TextFiles/'

# iterate over files in
# that directory
k = 2
for filename in os.listdir(directory):
    f = os.path.join(directory, filename)
    file = open(f, 'rt')
    text = file.read()
    file.close()
    #operations on the text
    #making it lowercase
    text = text.lower()

    #finding out no of words in sentence
    words= nltk.word_tokenize(text)
    len_wor= len(words)
    sentences= nltk.sent_tokenize(text)
    len_sen= len(sentences)
    avg_len_sen = 0.0
    avg_len_sen = len_wor/len_sen

    #counting syllables in a word
    com_words = 0.0
    tot_no_words = 0.0
    words = re.split(r'\W+', text)
    for word in words:
        tot_no_words = tot_no_words + 1
        syllable_len = len(p.positions(word))
        if syllable_len > 1:
            com_words = com_words + 1
    
    avg_com_words = (com_words/len_sen)
    fog_index = (avg_com_words + avg_len_sen) * 0.4
    #removing asciil characters
    text = text.encode('ascii', 'ignore').decode()
    #removing stopwords from nltk library
    stop_words = stopwords.words('english')
    text = ' '.join([word for word in text.split(' ') if word not in stop_words])
    #removing mentions such @who from text
    text = re.sub("@\S+", " ", text)
    #removing links from text
    text = re.sub("https*\S+", " ", text)
    #removing # from text
    text = re.sub("#\S+", " ", text)
    #removing apostrofes and next letter
    text = re.sub("\'\w+", '', text) 
    #removing punctuation
    text = re.sub('[%s]' % re.escape(string.punctuation), ' ', text)
    #removing numbers
    text = re.sub(r'\w*\d+\w*', '', text)
    #removing overspaces
    text = re.sub('\s{2,}', " ", text)

    #calculating each score
    cleaned_words = 0.0
    positive_score = 0.0
    negative_score = 0.0
    subjective_score = 0.0
    polarity_score = 0.0

    words = re.split(r'\W+', text)
    with open(r'C:/Users/Atharav Jadhav/source/repos/PythonApplication1/positive_words.txt', 'r') as fame:
        content = fame.read()
    # Iterate list to find each word
    for word in words:
        cleaned_words = cleaned_words + 1
        if word in content:
            positive_score = positive_score + 1;
     
    
    with open(r'C:/Users/Atharav Jadhav/source/repos/PythonApplication1/negative_words.txt', 'r') as fime:
        content = fime.read()
    # Iterate list to find each word
    for word in words:
        if word in content:
            negative_score = negative_score + 1;
    
    
    polarity_score = (positive_score-negative_score)/((positive_score+negative_score)+0.00001)
    subjective_score = (positive_score+negative_score)/((tot_no_words)+0.00001)
    
    
    fname = "Output.xlsx"
    wb = openpyxl.load_workbook(fname)
    ws = wb['Sheet1']
    ws.cell(k,3).value = positive_score
    ws.cell(k,4).value = negative_score
    ws.cell(k,5).value = polarity_score
    ws.cell(k,6).value = subjective_score
    ws.cell(k,7).value = avg_len_sen
    ws.cell(k,8).value = avg_com_words
    ws.cell(k,9).value = fog_index
    ws.cell(k,10).value = avg_len_sen
    ws.cell(k,11).value = com_words
    ws.cell(k,12).value = cleaned_words
    wb.save('Output.xlsx')
    k = k+1
    #to save our processed text list
    save_path = 'C:/Users/Atharav Jadhav/source/repos/PythonApplication1/ProcessedFiles/'

    completeName = os.path.join(save_path, f)

    file = open(completeName, 'w')
    file.writelines(text)
    file.close()

#processing of files is done and stored into a folder.


